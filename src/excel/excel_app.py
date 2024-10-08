import os
import openpyxl
import xlsxwriter
path = r"/Users/peng.guo/Documents/code/python-app/src/excel"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('templ.xlsx')	# 返回一个workbook数据类型的值
print(workbook.sheetnames)	# 打印Excel表中的所有表
sheet = workbook['Sheet1']  # 获取指定sheet表
print(sheet)
sheet = workbook.active     # 获取活动表
print(sheet)
print(sheet.dimensions)     # 获取表格的尺寸大小
cell1 = sheet['A1']         # 获取A1单元格的数据
cell2 = sheet['B1']         # 获取B7单元格的数据
# cell2 = sheet['B7'].value		# 另一种写法

# 正确示范
# cell1.value获取单元格A1中的值
# cell2.value获取单元格B7中的值
print(cell1.value,cell2.value)


cell1 = sheet.cell(row=1,column=1)         # 获取第1行第1列的数据
cell2 = sheet.cell(row=2,column=2)         # 获取第3行第4的数据

# 正确示范
# cell1.value获取单元格A1中的值
# cell2.value获取单元格B7中的值
print(cell1.value,cell2.value)  # 姓名 41


#获取workbook中所有的表格
sheets = workbook.get_sheet_names()
#循环遍历所有sheet
for i in range(len(sheets)):
    sheet= workbook.get_sheet_by_name(sheets[i])

    print('\n\n第'+str(i+1)+'个sheet: ' + sheet.title+'->>>')

    for r in range(1,sheet.max_row+1):
        if r == 1:
            print('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))
        else:
            print(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) for c in range(1,sheet.max_column+1)] ))

# 创建一个名为【demo.xlsx】工作簿；
workbook2 = xlsxwriter.Workbook("demo.xlsx")

# 创建一个名为【2018年销售量】工作表；
worksheet = workbook2.add_worksheet("2018年销售量")

# 使用write_row方法，为【2018年销售量】工作表，添加一个表头；
headings = ['产品', '销量', "单价"]
worksheet.write_row('A1', headings)
# 使用write方法，在【2018年销售量】工作表中插入一条数据；
# write语法格式：worksheet.write(行,列,数据)

data = ["苹果", 500, 8.9]
for i in range(len(headings)):
    worksheet.write(1, i, data[i])

workbook2.close()
